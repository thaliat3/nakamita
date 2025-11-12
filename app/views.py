from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt, ensure_csrf_cookie
from django.views.decorators.http import require_http_methods
from django.urls import reverse
from .models import Empleado, TipoAsistencia, RegistroAsistencia, DispositivoEmpleado, ActividadProyecto
from .services import AsistenciaService, ReporteService
from .qr_service import QRService
from .utils import obtener_fecha_hora_actual
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import json


def es_staff(user):
    """
    Verifica si el usuario es staff (administrador).
    
    Args:
        user: Usuario a verificar
        
    Returns:
        bool: True si es staff y está autenticado
    """
    return user.is_authenticated and user.is_staff

@user_passes_test(es_staff)
def pagina_descarga_excel(request):
    """
    Página para descargar reportes de Excel.
    Solo accesible para usuarios staff.
    """
    return render(request, 'pagina_descarga_excel.html')


@user_passes_test(es_staff)
def exportar_resumen_excel(request):
    """
    Exporta un resumen diario de asistencia en formato Excel.
    Incluye Proyecto y Actividad (si existen) para la ENTRADA de ese día.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen Diario"

    # Encabezados (agregamos Proyecto y Actividad)
    encabezados = [
        "Empleado", "Fecha", "Proyecto", "Actividad", "Tiempo de Almuerzo",
        "Horas por Comisión", "Horas por Permiso (Otros)",
        "Horas Trabajadas Totales"
    ]
    ws.append(encabezados)

    # Obtener datos usando el servicio
    datos_diarios = ReporteService.obtener_datos_resumen()

    from .models import ActividadProyecto
    from django.utils import timezone

    for (id_empleado, fecha), data in datos_diarios.items():
        empleado = data["empleado"]
        horas = ReporteService.calcular_horas_empleado(data)

        # Buscar actividad del día
        actividad = ActividadProyecto.objects.filter(empleado=empleado, fecha=fecha).first()
        proyecto_txt = actividad.proyecto if actividad else ''
        actividad_txt = actividad.actividad if actividad else ''

        ws.append([
            empleado.nombre_completo,
            fecha.strftime("%Y-%m-%d"),
            proyecto_txt,
            actividad_txt,
            horas['almuerzo'],
            horas['comision'],
            horas['permiso'],
            horas['trabajadas']
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Aplicar estilo al encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Crear tabla solo si hay datos (al menos 1 fila de datos)
    if ws.max_row > 1:
        tabla = Table(
            displayName="ResumenAsistencia",
            ref=f"A1:H{ws.max_row}"
        )
        style = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        tabla.tableStyleInfo = style
        ws.add_table(tabla)

    # Enviar archivo
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=resumen_asistencia.xlsx'
    wb.save(response)
    return response

@user_passes_test(es_staff)
def exportar_asistencia_excel(request):
    """
    Exporta todos los registros de asistencia en formato Excel.
    Incluye información detallada de cada registro y agrega una hoja "Actividades".
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    # Encabezados del excel
    encabezados = ["Empleado", "Tipo de Asistencia", "Fecha", "Hora", "Descripción", "ID Dispositivo"]
    ws.append(encabezados)

    # Registros usando el modelo
    registros = RegistroAsistencia.objects.select_related('empleado', 'tipo') \
        .order_by('-fecha_registro', '-hora_registro')
    
    for reg in registros:
        fila = [
            reg.empleado.nombre_completo,
            reg.tipo.nombre_asistencia,
            reg.fecha_registro.strftime('%Y-%m-%d'),
            reg.hora_registro.strftime('%H:%M:%S'),
            reg.descripcion or '',
            reg.fingerprint or '',
        ]
        ws.append(fila)

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Estilo encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Tabla solo si hay datos (al menos 1 fila de datos)
    if ws.max_row > 1:
        tabla = Table(
            displayName="RegistroAsistencia",
            ref=f"A1:F{ws.max_row}"
        )
        style = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        tabla.tableStyleInfo = style
        ws.add_table(tabla)

    # Hoja adicional: Actividades
    from .models import ActividadProyecto
    ws2 = wb.create_sheet(title="Actividades")
    ws2.append(["Empleado", "Fecha", "Proyecto", "Actividad", "Hora"])

    for ap in ActividadProyecto.objects.select_related('empleado').all().order_by('-fecha', '-hora'):
        ws2.append([
            ap.empleado.nombre_completo,
            ap.fecha.strftime('%Y-%m-%d'),
            ap.proyecto,
            ap.actividad,
            ap.hora.strftime('%H:%M:%S'),
        ])

    # Ajustar ancho y tabla hoja 2
    for col in ws2.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws2.column_dimensions[col[0].column_letter].width = max_length + 2

    if ws2.max_row > 1:
        tabla2 = Table(
            displayName="TablaActividades",
            ref=f"A1:E{ws2.max_row}"
        )
        style2 = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        tabla2.tableStyleInfo = style2
        ws2.add_table(tabla2)

    # Respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=registro_asistencia.xlsx'
    wb.save(response)
    return response



def pagina_principal(request):
    """
    Página principal con opciones de acceso.
    """
    return render(request, 'pagina_principal.html')


def escanear_qr(request):
    """
    Página para escanear código QR.
    """
    return render(request, 'escanear_qr.html')


@ensure_csrf_cookie
def identificar_dispositivo(request):
    """
    Página de QR general: identifica por fingerprint. Si ya está vinculado, redirige directo al formulario.
    Si no, muestra selector de empleado para vincular el dispositivo.
    """
    empleados = Empleado.objects.order_by('apellidos', 'nombres')
    return render(request, 'identificar.html', { 'empleados': empleados })


def registrar_asistencia_qr(request, codigo_qr):
    """
    Vista para registrar asistencia usando código QR.
    Detecta automáticamente al empleado.
    """
    empleado = Empleado.buscar_por_codigo_qr(codigo_qr)
    
    if not empleado:
        messages.error(request, 'Código QR no válido o empleado no encontrado.')
        return render(request, 'error_qr.html')
    
    tipos_evento = TipoAsistencia.objects.all()

    if request.method == 'POST':
        tipo_id = request.POST.get('tipo_evento')
        descripcion = request.POST.get('descripcion') or ''
        fingerprint = request.POST.get('fingerprint')

        # Validar datos requeridos
        if not tipo_id:
            messages.error(request, 'Debe seleccionar un tipo de asistencia.')
            return render(request, 'formulario_qr.html', {
                'empleado': empleado,
                'tipos_evento': tipos_evento
            })

        # Usar el servicio para crear el registro
        success, message, registro = AsistenciaService.crear_registro_asistencia(
            empleado.id_empleado, tipo_id, descripcion, fingerprint
        )

        if success:
            # Si es ENTRADA y aún no registra actividad hoy, ir a Control de Actividades (una sola vez)
            try:
                if registro.tipo.nombre_asistencia == 'Entrada':
                    from django.utils import timezone
                    hoy = timezone.localtime().date()
                    if not ActividadProyecto.objects.filter(empleado=registro.empleado, fecha=hoy).exists():
                        return redirect('control_actividades', empleado_id=registro.empleado.id_empleado)
            except Exception:
                pass

            messages.success(request, message)
            fecha, hora = obtener_fecha_hora_actual()
            return render(request, 'asistencia_exitosa.html', {
                'fecha': fecha,
                'hora': hora,
                'empleado': registro.empleado
            })
        else:
            messages.error(request, message)

    return render(request, 'formulario_qr.html', {
        'empleado': empleado,
        'tipos_evento': tipos_evento
    })


def registrar_asistencia_auto(request, empleado_id):
    """
    Registro usando identificación automática por fingerprint (QR general).
    Primera vez: se vincula en identificar_dispositivo.
    """
    empleado = get_object_or_404(Empleado, id_empleado=empleado_id)
    tipos_evento = TipoAsistencia.objects.all()

    if request.method == 'POST':
        tipo_id = request.POST.get('tipo_evento')
        descripcion = request.POST.get('descripcion') or ''
        fingerprint = request.POST.get('fingerprint')

        if not tipo_id:
            messages.error(request, 'Debe seleccionar un tipo de asistencia.')
            return render(request, 'formulario_qr.html', {
                'empleado': empleado,
                'tipos_evento': tipos_evento
            })

        success, message, registro = AsistenciaService.crear_registro_asistencia(
            empleado.id_empleado, tipo_id, descripcion, fingerprint
        )

        if success:
            # Si es ENTRADA y aún no registra actividad hoy, ir a Control de Actividades (una sola vez)
            try:
                if registro.tipo.nombre_asistencia == 'Entrada':
                    from django.utils import timezone
                    hoy = timezone.localtime().date()
                    if not ActividadProyecto.objects.filter(empleado=registro.empleado, fecha=hoy).exists():
                        return redirect('control_actividades', empleado_id=registro.empleado.id_empleado)
            except Exception:
                pass

            messages.success(request, message)
            fecha, hora = obtener_fecha_hora_actual()
            return render(request, 'asistencia_exitosa.html', {
                'fecha': fecha,
                'hora': hora,
                'empleado': registro.empleado
            })
        else:
            messages.error(request, message)

    return render(request, 'formulario_qr.html', {
        'empleado': empleado,
        'tipos_evento': tipos_evento
    })


@require_http_methods(["POST", "OPTIONS"])
def api_buscar_empleado_qr(request):
    """
    API para buscar empleado por código QR.
    """
    # Responder preflight/local OPTIONS
    if request.method == 'OPTIONS':
        return JsonResponse({'success': True})
    try:
        data = json.loads(request.body)
        codigo_qr = data.get('codigo_qr')
        
        if not codigo_qr:
            return JsonResponse({'success': False, 'error': 'Código QR requerido'}, status=400)
        
        resultado = QRService.buscar_empleado_por_qr(codigo_qr)
        status_code = 200 if resultado.get('success') else 404
        return JsonResponse(resultado, status=status_code)
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error del servidor: {str(e)}'}, status=500)


@require_http_methods(["POST", "OPTIONS"])
def api_identificar_por_fingerprint(request):
    """
    Identifica empleado por fingerprint del dispositivo.
    """
    if request.method == 'OPTIONS':
        return JsonResponse({'success': True})
    try:
        data = json.loads(request.body)
        fingerprint = data.get('fingerprint')
        if not fingerprint:
            return JsonResponse({'success': False, 'error': 'Fingerprint requerido'}, status=400)
        empleado = DispositivoEmpleado.obtener_empleado_por_fingerprint(fingerprint)
        if empleado:
            return JsonResponse({
                'success': True,
                'empleado': {
                    'id': empleado.id_empleado,
                    'nombres': empleado.nombres,
                    'apellidos': empleado.apellidos,
                    'nombre_completo': empleado.nombre_completo,
                    'dni': empleado.dni,
                }
            })
        else:
            return JsonResponse({'success': False, 'error': 'Dispositivo no vinculado a un empleado'}, status=404)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error del servidor: {str(e)}'}, status=500)


@require_http_methods(["POST", "OPTIONS"])
def api_vincular_fingerprint(request):
    """
    Vincula el fingerprint al empleado seleccionado (primera vez).
    """
    if request.method == 'OPTIONS':
        return JsonResponse({'success': True})
    try:
        data = json.loads(request.body)
        empleado_id = data.get('empleado_id')
        fingerprint = data.get('fingerprint')
        if not empleado_id or not fingerprint:
            return JsonResponse({'success': False, 'error': 'Empleado y fingerprint requeridos'}, status=400)
        empleado = Empleado.objects.get(id_empleado=empleado_id)
        # Reasignación permitida: si el fingerprint existe con otro empleado, se actualiza al elegido
        DispositivoEmpleado.objects.update_or_create(
            fingerprint=fingerprint,
            defaults={'empleado': empleado}
        )
        return JsonResponse({'success': True, 'empleado_id': empleado.id_empleado}, status=201)
    except Empleado.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Empleado no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error del servidor: {str(e)}'}, status=500)


@require_http_methods(["POST", "OPTIONS"])
def api_desvincular_fingerprint(request):
    """
    Desvincula el fingerprint del dispositivo actual para permitir seleccionar de nuevo.
    """
    if request.method == 'OPTIONS':
        return JsonResponse({'success': True})
    try:
        data = json.loads(request.body)
        fingerprint = data.get('fingerprint')
        if not fingerprint:
            return JsonResponse({'success': False, 'error': 'Fingerprint requerido'}, status=400)
        borrados, detalle = DispositivoEmpleado.objects.filter(fingerprint=fingerprint).delete()
        return JsonResponse({'success': True, 'deleted': borrados})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error del servidor: {str(e)}'}, status=500)


def registrar_asistencia(request):
    """
    Vista tradicional para registrar la asistencia de un empleado.
    Mantenida para compatibilidad.
    """
    empleados = Empleado.objects.all()
    tipos_evento = TipoAsistencia.objects.all()

    if request.method == 'POST':
        empleado_id = request.POST.get('empleado')
        tipo_id = request.POST.get('tipo_evento')
        descripcion = request.POST.get('descripcion') or ''
        fingerprint = request.POST.get('fingerprint')

        # Validar datos requeridos
        if not empleado_id or not tipo_id:
            messages.error(request, 'Debe seleccionar un empleado y tipo de asistencia.')
            return render(request, 'formulario.html', {
                'empleados': empleados,
                'tipos_evento': tipos_evento
            })

        # Usar el servicio para crear el registro
        success, message, registro = AsistenciaService.crear_registro_asistencia(
            empleado_id, tipo_id, descripcion, fingerprint
        )

        if success:
            # Si es ENTRADA y aún no registra actividad hoy, ir a Control de Actividades (una sola vez)
            try:
                if registro.tipo.nombre_asistencia == 'Entrada':
                    from django.utils import timezone
                    hoy = timezone.localtime().date()
                    if not ActividadProyecto.objects.filter(empleado=registro.empleado, fecha=hoy).exists():
                        return redirect('control_actividades', empleado_id=registro.empleado.id_empleado)
            except Exception:
                pass

            messages.success(request, message)
            fecha, hora = obtener_fecha_hora_actual()
            return render(request, 'asistencia_exitosa.html', {
                'fecha': fecha,
                'hora': hora,
                'empleado': registro.empleado
            })
        else:
            messages.error(request, message)

    return render(request, 'formulario.html', {
        'empleados': empleados,
        'tipos_evento': tipos_evento
    })


def control_actividades(request, empleado_id):
    """
    Registrar proyecto y actividad SOLO una vez por día, tras registrar la ENTRADA.
    Si ya existe para hoy, se omite y se redirige.
    """
    from django.utils import timezone
    empleado = get_object_or_404(Empleado, id_empleado=empleado_id)
    hoy = timezone.localtime().date()

    # Si ya existe registro hoy, omitir esta pantalla
    if ActividadProyecto.objects.filter(empleado=empleado, fecha=hoy).exists() and request.method == 'GET':
        messages.info(request, 'La actividad del día ya fue registrada.')
        return redirect('identificar_dispositivo')

    if request.method == 'POST':
        proyecto = (request.POST.get('proyecto') or '').strip()
        actividad = (request.POST.get('actividad') or '').strip()

        if not proyecto or not actividad:
            messages.error(request, 'Debe completar Proyecto y Actividad.')
            return render(request, 'control_actividades.html', {
                'empleado': empleado,
                'proyecto': proyecto,
                'actividad': actividad,
            })

        # Guardar localmente (único por día)
        ActividadProyecto.objects.update_or_create(
            empleado=empleado,
            fecha=hoy,
            defaults={'proyecto': proyecto, 'actividad': actividad}
        )
        messages.success(request, 'Actividad registrada correctamente.')

        # Mostrar pantalla de registro exitoso con datos, como en las otras opciones
        from .utils import obtener_fecha_hora_actual
        fecha, hora = obtener_fecha_hora_actual()
        return render(request, 'asistencia_exitosa.html', {
            'fecha': fecha,
            'hora': hora,
            'empleado': empleado,
        })

    return render(request, 'control_actividades.html', {'empleado': empleado})
